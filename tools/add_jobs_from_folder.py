
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


HEADERS = ["enabled", "input_video", "n", "module", "template", "output_name", "placeholder_video"]
VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".webm", ".m4v", ".avi"}


@dataclass(frozen=True)
class AddJobsArgs:
    manifest: Path
    folder: Path
    recursive: bool
    enabled: bool
    n: int
    module: str
    template: str
    output_name_prefix: str
    placeholder_video: str
    skip_existing: bool


def parse_args() -> AddJobsArgs:
    p = argparse.ArgumentParser(description="Add all video files from a folder into jobs.xlsx (sheet 'jobs').")
    p.add_argument("--manifest", required=True, help="Path to jobs.xlsx")
    p.add_argument("--folder", required=True, help="Folder containing videos")
    p.add_argument("--recursive", action="store_true", help="Scan subfolders too")

    p.add_argument("--enabled", default="true", help="true/false (default true)")
    p.add_argument("--n", type=int, default=6, help="Default number of quiz items (default 6)")
    p.add_argument("--module", default="list_reveal", help="Default module (default list_reveal)")
    p.add_argument("--template", default="6_word_template", help="Default template name")
    p.add_argument("--output-name-prefix", default="", help="If set, output_name becomes <prefix>_<index>")
    p.add_argument("--placeholder-video", default="", help="Placeholder filename inside template")
    p.add_argument("--skip-existing", action="store_true", help="Skip files already present in input_video column")

    a = p.parse_args()

    enabled_str = str(a.enabled).strip().lower()
    enabled = enabled_str in ("1", "true", "yes", "y", "on")

    return AddJobsArgs(
        manifest=Path(a.manifest),
        folder=Path(a.folder),
        recursive=bool(a.recursive),
        enabled=enabled,
        n=int(a.n),
        module=str(a.module),
        template=str(a.template),
        output_name_prefix=str(a.output_name_prefix),
        placeholder_video=str(a.placeholder_video),
        skip_existing=bool(a.skip_existing),
    )


def ensure_workbook_and_jobs_sheet(manifest: Path) -> Tuple[Workbook, Worksheet]:
    """
    Ensures:
      - workbook exists (create if missing)
      - a single sheet named 'jobs' exists
      - row 1 contains the required headers in exact order
    """
    if manifest.exists():
        wb = load_workbook(manifest)
    else:
        manifest.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

    # Ensure ONLY 'jobs' is used as the active sheet
    if "jobs" in wb.sheetnames:
        ws = wb["jobs"]
    else:
        # If new workbook, default sheet exists; rename it to jobs
        if len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
            ws = wb.active
            ws.title = "jobs"
        else:
            ws = wb.create_sheet("jobs")

    # Write header row exactly as required
    for col, name in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=name)

    return wb, ws


def scan_videos(folder: Path, recursive: bool) -> List[Path]:
    if recursive:
        paths = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in VIDEO_EXTS]
    else:
        paths = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in VIDEO_EXTS]
    return sorted(paths)


def existing_input_videos(ws: Worksheet) -> set[str]:
    # input_video is column 2
    seen: set[str] = set()
    max_row = ws.max_row or 1
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip():
            seen.add(v.strip().lower())
    return seen


def append_jobs(
    ws: Worksheet,
    videos: List[Path],
    enabled: bool,
    n: int,
    module: str,
    template: str,
    output_name_prefix: str,
    placeholder_video: str,
    skip_existing: bool,
) -> int:
    seen = existing_input_videos(ws) if skip_existing else set()

    # Find first empty row
    row = (ws.max_row or 1) + 1
    added = 0
    idx = 0

    for vp in videos:
        vp_str = str(vp.resolve())

        if skip_existing and vp_str.strip().lower() in seen:
            continue

        # output_name behavior:
        # - if prefix empty: leave blank (user/batch runner will auto-name)
        # - else: prefix_0, prefix_1, ...
        out_name = ""
        if output_name_prefix.strip():
            out_name = f"{output_name_prefix}_{idx}"
            idx += 1

        ws.cell(row=row, column=1, value="true" if enabled else "false")
        ws.cell(row=row, column=2, value=vp_str)
        ws.cell(row=row, column=3, value=int(n))
        ws.cell(row=row, column=4, value=str(module))
        ws.cell(row=row, column=5, value=str(template))
        ws.cell(row=row, column=6, value=str(out_name))
        ws.cell(row=row, column=7, value=str(placeholder_video))

        row += 1
        added += 1

    return added


def safe_save_workbook(wb: Workbook, manifest: Path) -> Path:
    """
    If Excel is open, saving will fail. In that case save to *_AUTOSAVED.xlsx.
    """
    try:
        wb.save(manifest)
        return manifest
    except PermissionError:
        fallback = manifest.with_name(manifest.stem + "_AUTOSAVED.xlsx")
        wb.save(fallback)
        print("WARNING: Cannot write to jobs.xlsx (likely open in Excel).")
        print(f"Saved to: {fallback}")
        return fallback


def main() -> int:
    args = parse_args()

    if not args.folder.exists() or not args.folder.is_dir():
        print(f"ERROR: folder not found: {args.folder}")
        return 2

    wb, ws = ensure_workbook_and_jobs_sheet(args.manifest)
    videos = scan_videos(args.folder, args.recursive)

    if not videos:
        print(f"No video files found in: {args.folder}")
        return 0

    added = append_jobs(
        ws=ws,
        videos=videos,
        enabled=args.enabled,
        n=args.n,
        module=args.module,
        template=args.template,
        output_name_prefix=args.output_name_prefix,
        placeholder_video=args.placeholder_video,
        skip_existing=args.skip_existing,
    )

    saved_to = safe_save_workbook(wb, args.manifest)

    print(f"Saved: {saved_to}")
    print(f"Videos found: {len(videos)}")
    print(f"Rows added: {added}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
